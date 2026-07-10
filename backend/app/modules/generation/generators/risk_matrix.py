"""Generador .xlsx de la Matriz de riesgos y oportunidades (6.1.1 + Anexo A).

La estructura de la hoja es fija (este código). La IA solo aporta las filas y
los campos planos, ya validados por Pydantic. Las celdas sin dato del cliente
se marcan ``[PENDIENTE: ...]``; nunca se inventan.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.modules.generation.generators.base import (
    DocumentGenerator,
    ResolvedField,
    resolve_code,
    resolve_text,
)
from app.modules.generation.schemas import RiskEntry, RiskMatrixVariables

# Orden de columnas de la matriz: (clave del campo, encabezado, ancho).
_COLUMNS: list[tuple[str, str, int]] = [
    ("activity", "Actividad", 22),
    ("hazard", "Peligro / Amenaza", 28),
    ("risk_description", "Riesgo", 32),
    ("affected", "Afectados", 18),
    ("probability", "Probabilidad", 14),
    ("severity", "Severidad", 14),
    ("risk_level", "Nivel de riesgo", 14),
    ("existing_controls", "Controles existentes", 30),
    ("proposed_actions", "Acciones propuestas", 30),
    ("responsible", "Responsable", 18),
    ("residual_risk", "Riesgo residual", 16),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


class RiskMatrixGenerator(DocumentGenerator):
    template_version = "matriz-riesgos-xlsx-v2"
    engine = "xlsx"
    custom_fields = frozenset({"risks"})

    def _extra_pending(
        self, variables: BaseModel, required_fields: set[str] | None
    ) -> list[str]:
        assert isinstance(variables, RiskMatrixVariables)
        if not variables.risks:
            return ["risks"] if self._is_required("risks", required_fields) else []
        # Pendientes de CELDA: cada campo vacío de cada fila real. No son claves
        # de la checklist (no mueven la completitud), pero el cliente y el
        # auditor deben verlos: son las celdas [PENDIENTE: ...] del archivo.
        pending: list[str] = []
        for i, entry in enumerate(variables.risks):
            for key, _, _ in _COLUMNS:
                value = getattr(entry, key)
                if not (value.strip() if isinstance(value, str) else value):
                    pending.append(f"risks[{i}].{key}")
        return pending

    def _render(
        self,
        resolved: dict[str, ResolvedField],
        variables: BaseModel,
        document_code: str | None,
    ) -> bytes:
        assert isinstance(variables, RiskMatrixVariables)

        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz de Riesgos"

        # --- Cabecera identidad (tipo MT: "solo encabezado", sin firmas) --
        # El código es DATO por tenant (document_codes); para Felipe MT-04
        # está confirmado (2026-07-05) y viene del catálogo, no del código.
        ncols = len(_COLUMNS)
        ws.cell(row=1, column=1, value="MATRIZ DE RIESGOS Y OPORTUNIDADES").font = Font(
            bold=True, size=14
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(
            row=2, column=1, value=f"Código: {resolve_code(document_code)} · Revisión: 01"
        ).font = Font(bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=3, column=1, value=resolved["company_name"].value)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        ws.cell(row=4, column=1, value="NTC-ISO 21101 — numeral 6.1.1 y Anexo A").font = (
            Font(italic=True)
        )
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
        ws.cell(row=5, column=1, value=f"Alcance: {resolved['scope'].value}")
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=ncols)
        ws.cell(row=6, column=1, value=f"Metodología: {resolved['methodology'].value}")
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=ncols)

        # --- Encabezado de la tabla ------------------------------------
        header_row = 7
        for col, (_, title, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _WRAP
            cell.border = _BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        # --- Filas de riesgo -------------------------------------------
        rows = variables.risks or [RiskEntry()]  # fila vacía => todo [PENDIENTE]
        field_desc = {k: v.description or k for k, v in RiskEntry.model_fields.items()}
        for r, entry in enumerate(rows, start=header_row + 1):
            for col, (key, _, _) in enumerate(_COLUMNS, start=1):
                text, _ = resolve_text(getattr(entry, key), field_desc[key])
                cell = ws.cell(row=r, column=col, value=text)
                cell.alignment = _WRAP
                cell.border = _BORDER
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # --- Hoja de oportunidades -------------------------------------
        ws_opp = wb.create_sheet("Oportunidades")
        ws_opp.cell(row=1, column=1, value="OPORTUNIDADES DE MEJORA (6.1.1)").font = Font(
            bold=True, size=12
        )
        ws_opp.column_dimensions["A"].width = 80
        opp_field = resolved["opportunities"]
        opportunities = (
            opp_field.value if isinstance(opp_field.value, list) else [opp_field.value]
        )
        for i, opp in enumerate(opportunities, start=3):
            ws_opp.cell(row=i, column=1, value=f"• {opp}").alignment = _WRAP

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
