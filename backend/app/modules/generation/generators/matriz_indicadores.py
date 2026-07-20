"""Generador .xlsx de la Matriz de indicadores de desempeño (numeral 9.1).

Reutiliza los indicadores que el cliente ya declaró en sus objetivos de
seguridad (6.2): una fila por indicador con nombre, objetivo, fórmula, meta,
frecuencia y responsable. El aporte propio del 9.1 es la ficha de SEGUIMIENTO
(período → resultado → estado → análisis), plantilla FIJA que la empresa
diligencia en la operación y que alimenta la revisión por la dirección (9.3).

El .xlsm original del kit es un tablero con macros y 80 hojas: aquí se genera la
matriz auditable (base de indicadores + ficha de seguimiento), sin la
automatización, que es una herramienta y no un documento por tenant.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.modules.generation.generators.base import (
    DocumentGenerator,
    ResolvedField,
    resolve_code,
    resolve_text,
)
from app.modules.generation.schemas import MatrizIndicadoresVariables, ObjectiveEntry

# Columnas de la base de indicadores (clave del campo en ObjectiveEntry o fijo).
_COLUMNS: list[tuple[str, str, int]] = [
    ("indicator_name", "Nombre del indicador", 26),
    ("objective", "Objetivo de seguridad que mide", 40),
    ("formula", "Fórmula", 30),
    ("goal", "Meta", 20),
    ("frequency", "Frecuencia de medición", 18),
    ("responsible", "Responsable de la medición", 22),
]

# Fuente de datos: no es dato del cliente; el molde deja un valor neutro para
# que la empresa lo precise (nunca se inventa un registro específico).
_DATA_SOURCE = "Registros del SGSTA asociados al indicador"

# Bloque fijo explicativo (Guía para la elaboración de indicadores del kit).
_INTRO = (
    "Los indicadores permiten evaluar periódicamente el desempeño del SGSTA "
    "frente a los objetivos de seguridad. Cada indicador se mide con su "
    "fórmula y se compara con la meta: se considera que CUMPLE cuando el "
    "resultado alcanza o supera la meta, y NO CUMPLE en caso contrario. Los "
    "resultados y su análisis se revisan en la revisión por la dirección (9.3)."
)

# Columnas de la ficha de SEGUIMIENTO (plantilla fija que la empresa diligencia).
_TRACKING_COLUMNS: list[tuple[str, int]] = [
    ("Período", 16),
    ("Variable 1", 14),
    ("Variable 2", 14),
    ("Resultado", 14),
    ("Meta", 14),
    ("Estado", 16),
    ("Análisis de resultados", 34),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


class MatrizIndicadoresGenerator(DocumentGenerator):
    template_version = "matriz-indicadores-xlsx-v1"
    engine = "xlsx"
    custom_fields = frozenset({"indicators"})
    static_sections = (
        "Bloque explicativo de la guía de indicadores (cómo se mide cada "
        "indicador con su fórmula y se compara con la meta; estados Cumple / "
        "No cumple)",
        "Ficha de seguimiento del indicador (período, variables, resultado, "
        "meta, estado y análisis) como plantilla fija que se diligencia en la "
        "operación y alimenta la revisión por la dirección",
        "Columna Fuente de datos con el valor neutro del molde (registros del "
        "SGSTA asociados al indicador)",
    )

    def _extra_pending(
        self, variables: BaseModel, required_fields: set[str] | None
    ) -> list[str]:
        assert isinstance(variables, MatrizIndicadoresVariables)
        pending: list[str] = []
        if not variables.indicators:
            if self._is_required("indicators", required_fields):
                pending.append("indicators")
        else:
            for i, entry in enumerate(variables.indicators):
                # Los campos medulares del indicador; frecuencia/responsable
                # pueden faltar sin invalidar la matriz.
                for key in ("indicator_name", "objective", "formula", "goal"):
                    value = getattr(entry, key)
                    if not (value.strip() if isinstance(value, str) else value):
                        pending.append(f"indicators[{i}].{key}")
        return pending

    def _render(
        self,
        resolved: dict[str, ResolvedField],
        variables: BaseModel,
        document_code: str | None,
    ) -> bytes:
        assert isinstance(variables, MatrizIndicadoresVariables)

        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz de indicadores"

        ncols = len(_COLUMNS) + 2  # + No. + Fuente de datos
        ws.cell(
            row=1, column=1, value="MATRIZ DE INDICADORES DE DESEMPEÑO DEL SGSTA"
        ).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(
            row=2, column=1, value=f"Código: {resolve_code(document_code)} · Revisión: 01"
        ).font = Font(bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=3, column=1, value=resolved["company_name"].value)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        ws.cell(row=4, column=1, value="NTC-ISO 21101 — numeral 9.1").font = Font(
            italic=True
        )
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)

        # --- Bloque explicativo fijo (guía de indicadores) ---------------
        intro = ws.cell(row=6, column=1, value=_INTRO)
        intro.alignment = _WRAP
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=ncols)
        ws.row_dimensions[6].height = 60

        # --- Base de indicadores: una fila por indicador -----------------
        header_row = 8
        headers = [("No.", 6)] + [(t, w) for _, t, w in _COLUMNS] + [
            ("Fuente de datos", 26)
        ]
        for col, (title, width) in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _WRAP
            cell.border = _BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        rows = variables.indicators or [ObjectiveEntry()]  # vacía => [PENDIENTE]
        field_desc = {
            k: v.description or k for k, v in ObjectiveEntry.model_fields.items()
        }
        for r, entry in enumerate(rows, start=header_row + 1):
            n = ws.cell(row=r, column=1, value=r - header_row)
            n.alignment = _WRAP
            n.border = _BORDER
            for offset, (key, _, _) in enumerate(_COLUMNS, start=2):
                text, _ = resolve_text(getattr(entry, key), field_desc[key])
                cell = ws.cell(row=r, column=offset, value=text)
                cell.alignment = _WRAP
                cell.border = _BORDER
            src = ws.cell(row=r, column=len(_COLUMNS) + 2, value=_DATA_SOURCE)
            src.alignment = _WRAP
            src.border = _BORDER

        # --- Ficha de seguimiento del indicador (plantilla fija) ---------
        track_title_row = header_row + len(rows) + 2
        ws.cell(
            row=track_title_row,
            column=1,
            value="FICHA DE SEGUIMIENTO DEL INDICADOR (se diligencia en cada medición)",
        ).font = Font(bold=True)
        ws.merge_cells(
            start_row=track_title_row, start_column=1,
            end_row=track_title_row, end_column=len(_TRACKING_COLUMNS),
        )
        track_header = track_title_row + 1
        for col, (title, _width) in enumerate(_TRACKING_COLUMNS, start=1):
            cell = ws.cell(row=track_header, column=col, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _WRAP
            cell.border = _BORDER
        # Filas vacías de la plantilla (bordeadas) para diligenciar.
        for r in range(track_header + 1, track_header + 5):
            for col in range(1, len(_TRACKING_COLUMNS) + 1):
                ws.cell(row=r, column=col).border = _BORDER
        legend_row = track_header + 6
        ws.cell(
            row=legend_row,
            column=1,
            value="Estado: «Cumple» si el resultado alcanza o supera la meta; "
            "«No cumple» en caso contrario.",
        ).font = Font(italic=True)
        ws.merge_cells(
            start_row=legend_row, start_column=1,
            end_row=legend_row, end_column=len(_TRACKING_COLUMNS),
        )

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
