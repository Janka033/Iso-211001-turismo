"""Generador .xlsx de la Planificación y control operacional (numeral 8.1).

Calca el molde del kit MinCIT: una fila por actividad de aventura con las
normas aplicables (nacional / internacional), los aspectos a controlar y cómo
se controlan. La estructura es fija; la IA solo aporta las filas, ya
validadas por Pydantic, derivadas de los datos reales del cliente. Las
normas citadas deben venir del contexto: nunca de memoria del modelo.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.modules.generation.generators.base import (
    DocumentGenerator,
    ResolvedField,
    resolve_code,
    resolve_text,
)
from app.modules.generation.schemas import (
    ControlOperacionalVariables,
    OperationalControlEntry,
)

# Columnas del molde: (clave del campo, encabezado, ancho).
_COLUMNS: list[tuple[str, str, int]] = [
    ("activity", "Actividad de aventura", 24),
    ("national_standards", "Normas aplicables — nacional", 34),
    ("international_standards", "Normas aplicables — internacional", 34),
    ("aspects", "Aspectos a controlar", 44),
    ("controls", "Cómo se controlan", 44),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


class ControlOperacionalGenerator(DocumentGenerator):
    template_version = "control-operacional-xlsx-v1"
    engine = "xlsx"
    custom_fields = frozenset({"controls"})

    def _extra_pending(
        self, variables: BaseModel, required_fields: set[str] | None
    ) -> list[str]:
        assert isinstance(variables, ControlOperacionalVariables)
        pending: list[str] = []
        if not variables.controls:
            if self._is_required("controls", required_fields):
                pending.append("controls")
        else:
            # Pendientes de CELDA (patrón de las matrices): cada campo vacío
            # de cada fila real, visible para el cliente y el auditor.
            for i, entry in enumerate(variables.controls):
                for key, _, _ in _COLUMNS:
                    value = getattr(entry, key)
                    if not (value.strip() if isinstance(value, str) else value):
                        pending.append(f"controls[{i}].{key}")
        return pending

    def _render(
        self,
        resolved: dict[str, ResolvedField],
        variables: BaseModel,
        document_code: str | None,
    ) -> bytes:
        assert isinstance(variables, ControlOperacionalVariables)

        wb = Workbook()
        ws = wb.active
        ws.title = "Control operacional"

        ncols = len(_COLUMNS)
        ws.cell(row=1, column=1, value="PLANIFICACIÓN Y CONTROL OPERACIONAL").font = Font(
            bold=True, size=14
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(
            row=2, column=1, value=f"Código: {resolve_code(document_code)} · Revisión: 01"
        ).font = Font(bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=3, column=1, value=resolved["company_name"].value)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        ws.cell(row=4, column=1, value="NTC-ISO 21101 — numeral 8.1").font = Font(
            italic=True
        )
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
        ws.cell(row=5, column=1, value="CONTROLES OPERACIONALES").font = Font(bold=True)
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=ncols)

        header_row = 6
        for col, (_, title, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _WRAP
            cell.border = _BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        rows = variables.controls or [OperationalControlEntry()]  # vacía => [PENDIENTE]
        field_desc = {
            k: v.description or k
            for k, v in OperationalControlEntry.model_fields.items()
        }
        for r, entry in enumerate(rows, start=header_row + 1):
            for col, (key, _, _) in enumerate(_COLUMNS, start=1):
                text, _ = resolve_text(getattr(entry, key), field_desc[key])
                cell = ws.cell(row=r, column=col, value=text)
                cell.alignment = _WRAP
                cell.border = _BORDER
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
