"""Generador .xlsx de la Matriz de objetivos de seguridad (numeral 6.2).

Calca el molde del kit MinCIT: bloque fijo de PLANIFICACIÓN (qué se hará,
recursos, evaluación, cuándo) + tabla con la política como marco y una fila
por objetivo (origen, indicador, fórmula, frecuencia, responsable, meta).
Las columnas RESULTADO y ANÁLISIS se diligencian en el seguimiento (revisión
por la dirección): salen con un texto neutro, no son pendientes del cliente.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.modules.generation.generators.base import (
    DocumentGenerator,
    ResolvedField,
    resolve_code,
    resolve_text,
)
from app.modules.generation.schemas import MatrizObjetivosVariables, ObjectiveEntry

# Bloque de PLANIFICACIÓN del molde MinCIT (contenido normativo fijo).
_PLANNING: tuple[tuple[str, str], ...] = (
    (
        "¿Qué se hará para lograr los objetivos?",
        "Para lograr los objetivos se establecen indicadores de gestión que "
        "periódicamente permiten identificar si se va cumpliendo la meta de "
        "cada objetivo del SGSTA.",
    ),
    (
        "¿Qué recursos se requieren para cumplir los objetivos?",
        "Los recursos asignados para cumplir los objetivos del SGSTA son "
        "económicos (presupuesto), de tiempo, humanos y los demás necesarios "
        "para lograr los objetivos y metas trazadas.",
    ),
    (
        "¿Cómo se evaluarán los resultados?",
        "Cada objetivo del SGSTA tiene un peso porcentual que en su sumatoria "
        "da el 100% de la eficacia del sistema; con los indicadores se "
        "identifica el avance frente a la meta de cada objetivo.",
    ),
    (
        "¿Cuándo se completará o finalizará la medición?",
        "En cada revisión por la dirección se analizará si los objetivos del "
        "SGSTA finalizan, continúan o requieren cambios por el análisis del "
        "contexto interno y externo o del desempeño del sistema.",
    ),
)

# Columnas por objetivo (clave del campo, encabezado, ancho).
_COLUMNS: list[tuple[str, str, int]] = [
    ("origin", "Origen del objetivo", 30),
    ("objective", "Objetivo de seguridad de turismo de aventura", 42),
    ("indicator_name", "Nombre del indicador", 24),
    ("formula", "Fórmula", 30),
    ("frequency", "Frecuencia de medición", 18),
    ("responsible", "¿Quién será responsable?", 20),
    ("goal", "Meta", 22),
]

_FOLLOW_UP = "(se diligencia en el seguimiento)"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


class ObjetivosSeguridadGenerator(DocumentGenerator):
    template_version = "objetivos-seguridad-xlsx-v1"
    engine = "xlsx"
    custom_fields = frozenset({"objectives"})
    static_sections = (
        "Bloque de planificación para lograr los objetivos (qué se hará, "
        "recursos, cómo se evaluarán los resultados, cuándo finaliza la "
        "medición) con el texto del kit MinCIT",
        "Columnas Resultado y Análisis de resultados reservadas al "
        "seguimiento de la revisión por la dirección",
    )

    def _extra_pending(
        self, variables: BaseModel, required_fields: set[str] | None
    ) -> list[str]:
        assert isinstance(variables, MatrizObjetivosVariables)
        pending: list[str] = []
        if not variables.objectives:
            if self._is_required("objectives", required_fields):
                pending.append("objectives")
        else:
            for i, entry in enumerate(variables.objectives):
                for key, _, _ in _COLUMNS:
                    value = getattr(entry, key)
                    if not (value.strip() if isinstance(value, str) else value):
                        pending.append(f"objectives[{i}].{key}")
        return pending

    def _render(
        self,
        resolved: dict[str, ResolvedField],
        variables: BaseModel,
        document_code: str | None,
    ) -> bytes:
        assert isinstance(variables, MatrizObjetivosVariables)

        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz de objetivos"

        ncols = len(_COLUMNS) + 3  # + política, resultado, análisis
        ws.cell(
            row=1, column=1, value="MATRIZ DE OBJETIVOS DE SEGURIDAD DE TURISMO DE AVENTURA"
        ).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(
            row=2, column=1, value=f"Código: {resolve_code(document_code)} · Revisión: 01"
        ).font = Font(bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=3, column=1, value=resolved["company_name"].value)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        ws.cell(row=4, column=1, value="NTC-ISO 21101 — numeral 6.2").font = Font(
            italic=True
        )
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)

        # --- Bloque fijo de planificación (molde MinCIT) -----------------
        ws.cell(
            row=6,
            column=1,
            value="PLANIFICACIÓN PARA LOGRAR LOS OBJETIVOS DE SEGURIDAD DE TURISMO DE AVENTURA",
        ).font = Font(bold=True)
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=ncols)
        for col, (question, answer) in enumerate(_PLANNING):
            start = 1 + col * 2
            q = ws.cell(row=7, column=start, value=question)
            q.fill = _HEADER_FILL
            q.font = _HEADER_FONT
            q.alignment = _WRAP
            q.border = _BORDER
            ws.merge_cells(start_row=7, start_column=start, end_row=7, end_column=start + 1)
            a = ws.cell(row=8, column=start, value=answer)
            a.alignment = _WRAP
            a.border = _BORDER
            ws.merge_cells(start_row=8, start_column=start, end_row=8, end_column=start + 1)
        ws.row_dimensions[8].height = 90

        # --- Encabezado de la tabla de objetivos -------------------------
        header_row = 10
        headers = (
            [("policy", "Política de seguridad de turismo de aventura", 34)]
            + _COLUMNS
            + [
                ("result", "Resultado", 18),
                ("analysis", "Análisis de resultados", 24),
            ]
        )
        for col, (_, title, width) in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _WRAP
            cell.border = _BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        # --- Filas: política como marco + un objetivo por fila -----------
        rows = variables.objectives or [ObjectiveEntry()]  # vacía => [PENDIENTE]
        field_desc = {
            k: v.description or k for k, v in ObjectiveEntry.model_fields.items()
        }
        first, last = header_row + 1, header_row + len(rows)
        policy_text, _ = resolve_text(
            variables.policy_statement,
            "Enunciado de la política de seguridad que enmarca los objetivos",
        )
        for r, entry in enumerate(rows, start=first):
            pol = ws.cell(row=r, column=1, value=policy_text if r == first else None)
            pol.alignment = _WRAP
            pol.border = _BORDER
            for offset, (key, _, _) in enumerate(_COLUMNS, start=2):
                text, _ = resolve_text(getattr(entry, key), field_desc[key])
                cell = ws.cell(row=r, column=offset, value=text)
                cell.alignment = _WRAP
                cell.border = _BORDER
            for col in (len(_COLUMNS) + 2, len(_COLUMNS) + 3):
                cell = ws.cell(row=r, column=col, value=_FOLLOW_UP)
                cell.alignment = _WRAP
                cell.border = _BORDER
        if last > first:
            ws.merge_cells(start_row=first, start_column=1, end_row=last, end_column=1)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
