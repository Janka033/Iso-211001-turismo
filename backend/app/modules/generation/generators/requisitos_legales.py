"""Generador .xlsx de la Matriz de requisitos legales (numeral 6.1.3).

Calca el MT-03 real de Felipe (TOURS AND ADVENTURES, aprobado 30/12/2025):
una fila por norma con tema, tipo, número, autoridad, disposición y acción de
cumplimiento. El CATÁLOGO LEGAL es contenido curado por el experto y vive
aquí TRANSCRITO EXACTO de su archivo — la IA NUNCA escribe una cita legal
(una ley mal citada invalida la auditoría). Cambios normativos entran por
``regulatory_changes`` con aprobación humana y una nueva versión de esta
plantilla. La evaluación de cumplimiento y la fecha de verificación son del
seguimiento vivo (capa APP), no de la generación.
"""

# ruff: noqa: E501 — las disposiciones y acciones del catálogo son datos
# transcritos EXACTOS del MT-03 de Felipe; no se parten para no alterarlos.

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.modules.generation.generators.base import (
    DocumentGenerator,
    ResolvedField,
    resolve_code,
)
from app.modules.generation.schemas import MatrizRequisitosLegalesVariables


@dataclass(frozen=True)
class LegalNorm:
    tema: str
    tipo: str
    numero: str
    autoridad: str
    disposicion: str
    accion: str
    #: Slugs de actividad a los que aplica; vacío = aplica a todo prestador.
    activity_tags: tuple[str, ...] = ()


# Catálogo base curado por Felipe (MT-03). Transcripción EXACTA del archivo.
_LEGAL_BASE: tuple[LegalNorm, ...] = (
    LegalNorm(
        tema="Calidad Turística",
        tipo="Resolución",
        numero="0612 de 2024",
        autoridad="MinCIT",
        disposicion="Reglamenta el artículo 12 de la Ley 2068 de 2020, estableciendo niveles de calidad turística y requisitos de certificación en sostenibilidad y seguridad.",
        accion="Cumplir con los niveles de calidad, obtener certificaciones obligatorias en turismo de aventura y ajustar procedimientos internos.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Comercio Electrónico",
        tipo="Ley",
        numero="527 de 1999",
        autoridad="Congreso de la República",
        disposicion="Regula el comercio electrónico, firma digital y transacciones digitales.",
        accion="Cumplir con normativas de comercio electrónico y garantizar seguridad en transacciones digitales.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Navegación y Transporte",
        tipo="Ley",
        numero="769 de 2002",
        autoridad="Congreso de Colombia",
        disposicion="Código Nacional de Tránsito Terrestre que regula el transporte y circulación de vehículos.",
        accion="Verificar documentación y permisos de vehículos turísticos.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Prevención ESCNNA",
        tipo="Ley",
        numero="679 de 2001",
        autoridad="Congreso de Colombia",
        disposicion="Establece el marco legal para prevenir y sancionar la explotación sexual comercial de niños, niñas y adolescentes en el turismo.",
        accion="Implementar políticas internas, capacitar personal y reportar actividades sospechosas.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Prevención ESCNNA",
        tipo="Ley",
        numero="1336 de 2009",
        autoridad="Congreso de Colombia",
        disposicion="Refuerza las medidas de control y sanciones para prevenir la explotación sexual comercial en el sector turístico.",
        accion="Capacitar personal, fortalecer controles y cumplir auditorías del MinCIT.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Prevención ESCNNA",
        tipo="Resolución",
        numero="3840 de 2009",
        autoridad="MinCIT",
        disposicion="Crea el Código de Conducta en el sector turístico para la prevención de la ESCNNA.",
        accion="Adoptar el código de conducta, incluir cláusulas en contratos y publicidad, y establecer mecanismos de denuncia.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Protección de Datos",
        tipo="Decreto",
        numero="1377 de 2013",
        autoridad="Presidencia - MinComercio - MinTIC",
        disposicion="Reglamenta la Ley 1581 de 2012 sobre protección de datos personales.",
        accion="Implementar medidas de seguridad en el manejo de datos personales y obtener consentimiento para su uso.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Protección de Datos",
        tipo="Ley",
        numero="1581 de 2012",
        autoridad="Congreso de Colombia",
        disposicion="Regula la recolección, almacenamiento, uso y eliminación de datos personales.",
        accion="Establecer políticas de privacidad y procedimientos para el manejo de datos personales.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Seguridad",
        tipo="Decreto",
        numero="945 de 2014",
        autoridad="Presidencia de la República - MinCIT",
        disposicion="Crea el Consejo Nacional de Seguridad Turística y exige implementación de protocolos de seguridad.",
        accion="Implementar protocolos de seguridad y capacitar al personal en gestión de emergencias.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Seguridad Contra Incendios",
        tipo="Ley",
        numero="9 de 1979",
        autoridad="Congreso de Colombia",
        disposicion="Requiere concepto técnico de bomberos para establecimientos después de su constitución y regula seguridad en evacuación y señalización.",
        accion="Solicitar concepto técnico de bomberos, implementar señalización y plan de evacuación.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Seguridad y Salud en el Trabajo",
        tipo="Resolución",
        numero="1111 de 2017",
        autoridad="Ministerio de Trabajo",
        disposicion="Define los estándares mínimos del SG-SST.",
        accion="Implementar un SG-SST y garantizar condiciones seguras.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Seguridad y Salud en el Trabajo",
        tipo="Resolución",
        numero="2400 de 1979",
        autoridad="Min. de Trabajo y Seguridad Social",
        disposicion="Regula higiene, vivienda y seguridad en establecimientos de trabajo.",
        accion="Cumplir requisitos de higiene y seguridad en instalaciones físicas.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Seguridad y Salud en el Trabajo",
        tipo="Ley",
        numero="9 de 1979",
        autoridad="Congreso de Colombia",
        disposicion="Código Sanitario Nacional, seguridad industrial y comités de seguridad.",
        accion="Adoptar medidas de higiene y conformar comité de seguridad.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Seguridad y Salud en el Trabajo",
        tipo="Decreto",
        numero="171 de 2016",
        autoridad="Ministerio de Trabajo",
        disposicion="Obliga a sustituir el programa de Salud Ocupacional por el SG-SST.",
        accion="Implementar completamente el SG-SST en la empresa.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Seguridad y Salud en el Trabajo",
        tipo="Decreto",
        numero="1072 de 2015",
        autoridad="Min. de Trabajo y Seguridad Social",
        disposicion="Decreto Único Reglamentario del Sector Trabajo.",
        accion="Aplicar las disposiciones relevantes del sector trabajo.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Seguridad y Salud en el Trabajo",
        tipo="Ley",
        numero="1295 de 1994",
        autoridad="Congreso de Colombia",
        disposicion="Regula la organización y administración del Sistema General de Riesgos Laborales.",
        accion="Garantizar afiliaciones a riesgos laborales y estándares de seguridad.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Seguridad y Salud en el Trabajo",
        tipo="Resolución",
        numero="0312 de 2019",
        autoridad="Ministerio de Trabajo",
        disposicion="Define estándares mínimos del SG-SST y complementa la Resolución 1111 de 2017.",
        accion="Cumplir con los requisitos del SG-SST, especialmente en turismo de aventura.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Turismo",
        tipo="Ley",
        numero="300 de 1996",
        autoridad="Congreso de Colombia",
        disposicion="Ley General de Turismo que regula la actividad turística y establece el Registro Nacional de Turismo.",
        accion="Inscripción en el RNT y cumplimiento de estándares de operación.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Turismo",
        tipo="Ley",
        numero="1101 de 2006",
        autoridad="Congreso de Colombia",
        disposicion="Modifica la Ley 300 en lo relacionado con la contribución parafiscal para promoción del turismo.",
        accion="Cumplir con la contribución parafiscal.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Turismo",
        tipo="Ley",
        numero="1558 de 2012",
        autoridad="Congreso de Colombia",
        disposicion="Modifica la Ley 300 y establece el cumplimiento obligatorio de Normas Técnicas Sectoriales (NTS).",
        accion="Implementación de NTS aplicables a cada actividad.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Turismo",
        tipo="Decreto",
        numero="1074 de 2015",
        autoridad="Presidencia de la República",
        disposicion="Reglamenta el sector comercio y turismo, incluyendo inscripción en el RNT. Este decreto compila las normas del sector Comercio, Industria y Turismo, derogando varios decretos anteriores, incluyendo los Decretos 503 de 1997, 504 de 1997, 2074 de 2003, 053 de 2002 y 2438 de 2010.",
        accion="Verificación de inscripción y cumplimiento de requisitos.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Turismo y Regulación",
        tipo="Ley",
        numero="2068 de 2020",
        autoridad="Congreso de Colombia",
        disposicion="Modifica la Ley General de Turismo, regula la sostenibilidad, calidad, plataformas digitales, sanciones, incentivos y contribuciones fiscales.",
        accion="Cumplir con lineamientos de sostenibilidad, regulación de plataformas, calidad en servicios y normativas de contribución parafiscal.",
        activity_tags=(),
    ),
    LegalNorm(
        tema="Vehículos",
        tipo="Resolución",
        numero="3124 de 2014",
        autoridad="MinTransporte",
        disposicion="Regula la circulación de cuatrimotos en vías públicas y privadas.",
        accion="Garantizar documentación, placas y seguro obligatorio para cuatrimotos turísticos.",
        activity_tags=("atv",),
    ),
)

_COLUMNS: list[tuple[str, int]] = [
    ("Tema", 24),
    ("Tipo", 12),
    ("Número y fecha", 16),
    ("Autoridad", 26),
    ("Disposición", 52),
    ("Acción de cumplimiento", 46),
    ("Evaluación de cumplimiento", 22),
    ("Fecha de verificación", 18),
]

_FOLLOW_UP = "(se diligencia en el seguimiento)"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


def applicable_norms(activities: list[str]) -> list[LegalNorm]:
    """Filtro DETERMINISTA del catálogo: las normas etiquetadas por actividad
    solo entran si el tenant ofrece esa actividad; el resto aplica siempre."""
    acts = {a.strip().lower() for a in activities}
    return [
        n
        for n in _LEGAL_BASE
        if not n.activity_tags or any(t in acts for t in n.activity_tags)
    ]


class RequisitosLegalesGenerator(DocumentGenerator):
    template_version = "requisitos-legales-xlsx-v1"
    engine = "xlsx"
    static_sections = (
        "Catálogo legal base curado por el experto (23 normas: turismo y RNT, "
        "SG-SST, prevención ESCNNA, protección de datos, seguridad, "
        "incendios, tránsito y calidad turística) transcrito exacto del "
        "MT-03; la fila de cuatrimotos solo aplica a tenants con ATV",
        "Columnas Evaluación de cumplimiento y Fecha de verificación "
        "reservadas al seguimiento vivo",
    )

    def _render(
        self,
        resolved: dict[str, ResolvedField],
        variables: BaseModel,
        document_code: str | None,
    ) -> bytes:
        assert isinstance(variables, MatrizRequisitosLegalesVariables)

        wb = Workbook()
        ws = wb.active
        ws.title = "MT-03"

        ncols = len(_COLUMNS)
        ws.cell(row=1, column=1, value="MATRIZ DE REQUISITOS LEGALES").font = Font(
            bold=True, size=14
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(
            row=2, column=1, value=f"Código: {resolve_code(document_code)} · Revisión: 01"
        ).font = Font(bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=3, column=1, value=resolved["company_name"].value)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        ws.cell(row=4, column=1, value="NTC-ISO 21101 — numeral 6.1.3").font = Font(
            italic=True
        )
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)

        header_row = 6
        for col, (title, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _WRAP
            cell.border = _BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        norms = applicable_norms(variables.activities)
        for r, norm in enumerate(norms, start=header_row + 1):
            values = (
                norm.tema,
                norm.tipo,
                norm.numero,
                norm.autoridad,
                norm.disposicion,
                norm.accion,
                _FOLLOW_UP,
                _FOLLOW_UP,
            )
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col, value=value)
                cell.alignment = _WRAP
                cell.border = _BORDER
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
