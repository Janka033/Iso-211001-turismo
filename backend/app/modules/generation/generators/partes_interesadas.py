"""Generador .xlsx de la Matriz de partes interesadas (numeral 4.2).

Calca la matriz del kit MinCIT: una fila por parte interesada con necesidades,
expectativas, evaluación de cumplimiento (C/CP/NC) y plan de acción. La
estructura de la hoja es fija (este código); la IA solo aporta las filas, ya
validadas por Pydantic. Las celdas sin dato del cliente se marcan
``[PENDIENTE: ...]``; nunca se inventan.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.modules.generation.generators.base import (
    DocumentGenerator,
    ResolvedField,
    resolve_code,
    resolve_text,
)
from app.modules.generation.schemas import (
    MatrizPartesInteresadasVariables,
    StakeholderEntry,
)

# Orden de columnas de la matriz del kit: (clave del campo, encabezado, ancho).
_COLUMNS: list[tuple[str, str, int]] = [
    ("stakeholder", "Parte interesada perteneciente al SGSTA", 34),
    ("needs", "Necesidad(es) de la parte interesada", 36),
    ("expectations", "Expectativa(s) de la parte interesada", 36),
    ("compliance", "¿La organización cumple los requisitos? (C / CP / NC)", 16),
    ("observations", "Observaciones (aplica para CP y NC)", 30),
    ("actions", "Acciones a realizar (aplica para CP y NC)", 32),
    ("execution_date", "Fecha para la ejecución", 16),
    ("responsible", "Responsable de ejecución", 20),
    ("follow_up_date", "Fecha de seguimiento", 16),
    ("status", "Estado (abierta / cerrada)", 14),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


class PartesInteresadasGenerator(DocumentGenerator):
    template_version = "partes-interesadas-xlsx-v1"
    engine = "xlsx"
    custom_fields = frozenset({"stakeholders"})

    def _extra_pending(
        self, variables: BaseModel, required_fields: set[str] | None
    ) -> list[str]:
        assert isinstance(variables, MatrizPartesInteresadasVariables)
        pending: list[str] = []
        if not variables.stakeholders:
            if self._is_required("stakeholders", required_fields):
                pending.append("stakeholders")
        else:
            # Pendientes de CELDA (patrón de la matriz de riesgos): cada campo
            # vacío de cada fila real, visible para el cliente y el auditor.
            for i, entry in enumerate(variables.stakeholders):
                for key, _, _ in _COLUMNS:
                    value = getattr(entry, key)
                    if not (value.strip() if isinstance(value, str) else value):
                        pending.append(f"stakeholders[{i}].{key}")
        return pending

    def _render(
        self,
        resolved: dict[str, ResolvedField],
        variables: BaseModel,
        document_code: str | None,
    ) -> bytes:
        assert isinstance(variables, MatrizPartesInteresadasVariables)

        wb = Workbook()
        ws = wb.active
        ws.title = "Partes interesadas"

        # --- Cabecera identidad (tipo MT: "solo encabezado", sin firmas) --
        ncols = len(_COLUMNS)
        ws.cell(row=1, column=1, value="MATRIZ DE PARTES INTERESADAS").font = Font(
            bold=True, size=14
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(
            row=2, column=1, value=f"Código: {resolve_code(document_code)} · Revisión: 01"
        ).font = Font(bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=3, column=1, value=resolved["company_name"].value)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        ws.cell(row=4, column=1, value="NTC-ISO 21101 — numeral 4.2").font = Font(
            italic=True
        )
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
        ws.cell(
            row=5,
            column=1,
            value=f"Responsable de diligenciar: {resolved['elaborated_by'].value}",
        )
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=ncols)

        # --- Encabezado de la tabla ------------------------------------
        header_row = 6
        for col, (_, title, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _WRAP
            cell.border = _BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        # --- Filas ------------------------------------------------------
        rows = variables.stakeholders or [StakeholderEntry()]  # vacía => [PENDIENTE]
        field_desc = {
            k: v.description or k for k, v in StakeholderEntry.model_fields.items()
        }
        for r, entry in enumerate(rows, start=header_row + 1):
            for col, (key, _, _) in enumerate(_COLUMNS, start=1):
                text, _ = resolve_text(getattr(entry, key), field_desc[key])
                cell = ws.cell(row=r, column=col, value=text)
                cell.alignment = _WRAP
                cell.border = _BORDER
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
