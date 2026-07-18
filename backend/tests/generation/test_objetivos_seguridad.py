"""Tests de la Matriz de objetivos de seguridad (6.2): generador 4/8.

Sin preguntas nuevas de onboarding: su insumo (safety_objectives) se captura
en el paso de la política. Resultado/análisis son del seguimiento, no
pendientes del cliente.
"""

import io

from openpyxl import load_workbook

from app.modules.generation.generators.factory import get_spec
from app.modules.generation.generators.objetivos_seguridad import (
    ObjetivosSeguridadGenerator,
)
from app.modules.generation.schemas import MatrizObjetivosVariables, ObjectiveEntry


def _cells(content: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(content))
    values: list[str] = []
    for ws in wb:
        for row in ws.iter_rows():
            values.extend(str(c.value) for c in row if c.value is not None)
    return values


def test_objetivos_full_rows():
    variables = MatrizObjetivosVariables(
        company_name="Rafting Fonce SAS",
        policy_statement="La dirección prioriza la seguridad en toda la operación.",
        objectives=[
            ObjectiveEntry(
                origin="Riesgos y oportunidades",
                objective="Reducir los incidentes reportados en 20% a diciembre de 2026",
                indicator_name="Tasa de incidentes",
                formula="(incidentes del periodo / salidas del periodo) x 100",
                frequency="Mensual",
                responsible="Gerente",
                goal="-20% a diciembre de 2026",
            )
        ],
    )
    result = ObjetivosSeguridadGenerator().generate(variables)

    assert result.pending_fields == []
    txt = "\n".join(_cells(result.content))
    assert "MATRIZ DE OBJETIVOS DE SEGURIDAD" in txt
    assert "Reducir los incidentes reportados" in txt
    assert "Tasa de incidentes" in txt
    assert "La dirección prioriza la seguridad" in txt
    # Bloque fijo de planificación del molde y columnas de seguimiento.
    assert "¿Qué se hará para lograr los objetivos?" in txt
    assert "(se diligencia en el seguimiento)" in txt


def test_objetivos_empty_marks_pending():
    result = ObjetivosSeguridadGenerator().generate(MatrizObjetivosVariables())
    assert "objectives" in result.pending_fields
    assert "company_name" in result.pending_fields
    assert any("[PENDIENTE:" in v for v in _cells(result.content))


def test_objetivos_cell_pendings_reported():
    variables = MatrizObjetivosVariables(
        company_name="X SAS",
        policy_statement="Política.",
        objectives=[
            ObjectiveEntry(
                origin="Contexto",
                objective="Capacitar al 100% de los guías",
                # sin indicador/fórmula/frecuencia/responsable/meta
            )
        ],
    )
    result = ObjetivosSeguridadGenerator().generate(variables)
    assert "objectives[0].indicator_name" in result.pending_fields
    assert "objectives[0].goal" in result.pending_fields
    # El seguimiento NO es pendiente del cliente.
    assert not any("result" in p or "analysis" in p for p in result.pending_fields)


def test_objetivos_registered_in_factory():
    spec = get_spec("matriz_objetivos_seguridad")
    assert spec is not None
    assert spec.numeral == "6.2"
    assert spec.variables_model is MatrizObjetivosVariables
    assert spec.rag_numerales == ("6.2", "5.2")
    assert spec.generator.engine == "xlsx"
    assert spec.generator.static_sections
