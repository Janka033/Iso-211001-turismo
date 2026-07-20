"""Tests de la Matriz de indicadores de desempeño (9.1).

Reutiliza los indicadores de los objetivos de seguridad (6.2); la ficha de
seguimiento y el bloque explicativo son plantilla fija. La IA no inventa
indicadores.
"""

import io

from openpyxl import load_workbook

from app.modules.generation.generators.factory import get_spec
from app.modules.generation.generators.matriz_indicadores import (
    MatrizIndicadoresGenerator,
)
from app.modules.generation.schemas import MatrizIndicadoresVariables, ObjectiveEntry


def _text(content: bytes) -> str:
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    parts: list[str] = []
    for row in ws.iter_rows(values_only=True):
        parts.extend(str(c) for c in row if c is not None)
    return "\n".join(parts)


def _indicator(**kw) -> ObjectiveEntry:
    base = dict(
        indicator_name="Índice de accidentalidad",
        objective="Reducir los incidentes reportados 20% a diciembre 2026",
        formula="(N.º incidentes / N.º salidas) x 100",
        goal="≤ 2% a diciembre 2026",
        frequency="Mensual",
        responsible="Gerente",
    )
    base.update(kw)
    return ObjectiveEntry(**base)


def test_matriz_indicadores_full_document():
    variables = MatrizIndicadoresVariables(
        company_name="Rafting Fonce SAS",
        indicators=[_indicator(), _indicator(indicator_name="Cobertura de capacitación")],
    )
    result = MatrizIndicadoresGenerator().generate(variables)

    assert result.pending_fields == []
    txt = _text(result.content)
    # Datos del cliente (indicadores derivados de 6.2).
    assert "Índice de accidentalidad" in txt
    assert "Cobertura de capacitación" in txt
    assert "(N.º incidentes / N.º salidas) x 100" in txt
    # Contenido fijo del 9.1.
    assert "FICHA DE SEGUIMIENTO DEL INDICADOR" in txt
    assert "Cumple" in txt  # leyenda de estado
    assert "Registros del SGSTA" in txt  # fuente de datos neutra
    assert "9.1" in txt


def test_matriz_indicadores_empty_marks_pending():
    result = MatrizIndicadoresGenerator().generate(MatrizIndicadoresVariables())
    assert "indicators" in result.pending_fields
    txt = _text(result.content)
    assert "[PENDIENTE:" in txt
    # La ficha de seguimiento fija sale aunque no haya indicadores.
    assert "FICHA DE SEGUIMIENTO DEL INDICADOR" in txt


def test_matriz_indicadores_incomplete_indicator_reports_pending():
    variables = MatrizIndicadoresVariables(
        company_name="X SAS",
        indicators=[_indicator(formula=None, goal=None)],
    )
    result = MatrizIndicadoresGenerator().generate(variables)
    assert "indicators[0].formula" in result.pending_fields
    assert "indicators[0].goal" in result.pending_fields


def test_matriz_indicadores_registered_in_factory():
    spec = get_spec("matriz_indicadores")
    assert spec is not None
    assert spec.numeral == "9.1"
    assert spec.variables_model is MatrizIndicadoresVariables
    assert spec.rag_numerales == ("9.1", "6.2")
    assert spec.generator.engine == "xlsx"
