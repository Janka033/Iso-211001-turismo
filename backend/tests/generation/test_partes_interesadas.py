"""Tests de la Matriz de partes interesadas (4.2): generador 2/8 de la ruta.

Mismo patrón que la matriz de riesgos: filas estructuradas, pendientes de
celda visibles y registro en la factory.
"""

import io

from openpyxl import load_workbook

from app.modules.generation.generators.factory import get_spec
from app.modules.generation.generators.partes_interesadas import (
    PartesInteresadasGenerator,
)
from app.modules.generation.schemas import (
    MatrizPartesInteresadasVariables,
    StakeholderEntry,
)


def _cells(content: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(content))
    values: list[str] = []
    for ws in wb:
        for row in ws.iter_rows():
            values.extend(str(c.value) for c in row if c.value is not None)
    return values


def test_partes_interesadas_full_row():
    variables = MatrizPartesInteresadasVariables(
        company_name="Rafting Fonce SAS",
        elaborated_by="Gerente",
        stakeholders=[
            StakeholderEntry(
                stakeholder="Organizaciones gubernamentales — Alcaldía de San Gil",
                needs="RNT vigente y cumplimiento normativo",
                expectations="Operación segura y formalizada",
                compliance="C",
                observations="No aplica",
                actions="No aplica",
                execution_date="No aplica",
                responsible="Gerente",
                follow_up_date="2026-06-30",
                status="No aplica",
            )
        ],
    )
    result = PartesInteresadasGenerator().generate(variables)

    assert result.pending_fields == []
    txt = "\n".join(_cells(result.content))
    assert "MATRIZ DE PARTES INTERESADAS" in txt
    assert "Alcaldía de San Gil" in txt
    assert "RNT vigente y cumplimiento normativo" in txt
    assert "Responsable de diligenciar: Gerente" in txt
    assert "numeral 4.2" in txt


def test_partes_interesadas_empty_marks_pending():
    result = PartesInteresadasGenerator().generate(MatrizPartesInteresadasVariables())
    assert "stakeholders" in result.pending_fields
    assert "company_name" in result.pending_fields
    assert any("[PENDIENTE:" in v for v in _cells(result.content))


def test_partes_interesadas_cell_pendings_reported():
    """Una fila real con celdas vacías reporta cada pendiente de celda."""
    variables = MatrizPartesInteresadasVariables(
        company_name="X SAS",
        elaborated_by="Gerente",
        stakeholders=[
            StakeholderEntry(
                stakeholder="Participantes / clientes",
                needs="Actividad segura",
                # sin compliance ni plan: el cliente no lo declaró
            )
        ],
    )
    result = PartesInteresadasGenerator().generate(variables)
    assert "stakeholders[0].compliance" in result.pending_fields
    assert "stakeholders[0].actions" in result.pending_fields
    assert "stakeholders[0].expectations" in result.pending_fields


def test_partes_interesadas_registered_in_factory():
    spec = get_spec("matriz_partes_interesadas")
    assert spec is not None
    assert spec.numeral == "4.2"
    assert spec.variables_model is MatrizPartesInteresadasVariables
    assert spec.rag_numerales == ("4.2", "4.1")
    assert spec.generator.engine == "xlsx"
