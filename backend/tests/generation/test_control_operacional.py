"""Tests de la Planificación y control operacional (8.1): generador 7/8.

Una fila por actividad con normas aplicables, aspectos a controlar y cómo se
controlan; pendientes de celda visibles.
"""

import io

from openpyxl import load_workbook

from app.modules.generation.generators.control_operacional import (
    ControlOperacionalGenerator,
)
from app.modules.generation.generators.factory import get_spec
from app.modules.generation.schemas import (
    ControlOperacionalVariables,
    OperationalControlEntry,
)


def _cells(content: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(content))
    values: list[str] = []
    for ws in wb:
        for row in ws.iter_rows():
            values.extend(str(c.value) for c in row if c.value is not None)
    return values


def test_control_operacional_full_row():
    variables = ControlOperacionalVariables(
        company_name="Rafting Fonce SAS",
        controls=[
            OperationalControlEntry(
                activity="Rafting",
                national_standards="NTC-ISO 21101 (mencionada por el cliente)",
                international_standards="No identificadas por el cliente",
                aspects="Equipos (balsas, chalecos, cascos); competencias de los guías",
                controls="Inspección pre-operacional diaria; briefing de seguridad",
            )
        ],
    )
    result = ControlOperacionalGenerator().generate(variables)

    assert result.pending_fields == []
    txt = "\n".join(_cells(result.content))
    assert "PLANIFICACIÓN Y CONTROL OPERACIONAL" in txt
    assert "Rafting" in txt
    assert "Inspección pre-operacional diaria" in txt
    assert "numeral 8.1" in txt


def test_control_operacional_empty_marks_pending():
    result = ControlOperacionalGenerator().generate(ControlOperacionalVariables())
    assert "controls" in result.pending_fields
    assert "company_name" in result.pending_fields
    assert any("[PENDIENTE:" in v for v in _cells(result.content))


def test_control_operacional_cell_pendings_reported():
    variables = ControlOperacionalVariables(
        company_name="X SAS",
        controls=[
            OperationalControlEntry(
                activity="Rafting",
                aspects="Equipos",
                # sin normas ni controles
            )
        ],
    )
    result = ControlOperacionalGenerator().generate(variables)
    assert "controls[0].national_standards" in result.pending_fields
    assert "controls[0].controls" in result.pending_fields


def test_control_operacional_registered_in_factory():
    spec = get_spec("control_operacional")
    assert spec is not None
    assert spec.numeral == "8.1"
    assert spec.variables_model is ControlOperacionalVariables
    assert spec.rag_numerales == ("8.1", "A")
    assert spec.generator.engine == "xlsx"
