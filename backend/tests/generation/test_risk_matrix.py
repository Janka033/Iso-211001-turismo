"""Tests del generador .xlsx de la matriz de riesgos y de su flujo en el endpoint."""

import io

from openpyxl import load_workbook

from app.modules.generation import repository, service
from app.modules.generation.generators.risk_matrix import RiskMatrixGenerator
from app.modules.generation.schemas import RiskEntry, RiskMatrixVariables
from app.modules.inventory import service as inventory_service

TENANT_A = "11111111-1111-1111-1111-111111111111"


def _load(content: bytes):
    return load_workbook(io.BytesIO(content))


def test_full_matrix_no_pending():
    variables = RiskMatrixVariables(
        company_name="Aventura Andina SAS",
        scope="Rafting en el río Suárez.",
        methodology="Probabilidad × severidad, 3x3.",
        risks=[
            RiskEntry(
                activity="Rafting",
                hazard="Volcamiento de balsa",
                risk_description="Caída al agua y ahogamiento",
                affected="Turistas",
                probability="Media",
                severity="Grave",
                risk_level="Alto",
                existing_controls="Chalecos y cascos certificados",
                proposed_actions="Briefing obligatorio y guía por balsa",
                responsible="Coordinador de operaciones",
                residual_risk="Medio",
            )
        ],
        opportunities=["Certificar guías en rescate acuático"],
    )

    result = RiskMatrixGenerator().generate(variables, document_code='MT-04')

    assert result.pending_fields == []
    wb = _load(result.content)
    assert "Matriz de Riesgos" in wb.sheetnames
    assert "Oportunidades" in wb.sheetnames
    # La primera fila de datos (fila 8) lleva la actividad.
    ws = wb["Matriz de Riesgos"]
    assert ws.cell(row=8, column=1).value == "Rafting"
    # Sin placeholders en ninguna celda de datos.
    flat = [c.value for row in ws.iter_rows() for c in row if c.value]
    assert not any("[PENDIENTE" in str(v) for v in flat)


def test_empty_risks_marks_pending_and_renders_placeholder_row():
    variables = RiskMatrixVariables(company_name="Sin Datos SAS")
    result = RiskMatrixGenerator().generate(variables)

    # risks vacío => pendiente de alto nivel.
    assert "risks" in result.pending_fields
    assert "scope" in result.pending_fields
    assert "methodology" in result.pending_fields

    ws = _load(result.content)["Matriz de Riesgos"]
    # Fila de datos vacía => todas las celdas en [PENDIENTE].
    assert "[PENDIENTE:" in str(ws.cell(row=8, column=1).value)


def test_partial_row_only_missing_cells_pending():
    variables = RiskMatrixVariables(
        company_name="X SAS",
        scope="s",
        methodology="m",
        risks=[RiskEntry(activity="Canopy")],  # solo actividad
    )
    result = RiskMatrixGenerator().generate(variables)
    assert "risks" not in result.pending_fields  # hay al menos una fila
    # Los pendientes de celda se reportan uno a uno (visibles para el cliente).
    assert "risks[0].hazard" in result.pending_fields
    assert "risks[0].probability" in result.pending_fields
    assert "risks[0].activity" not in result.pending_fields

    ws = _load(result.content)["Matriz de Riesgos"]
    assert ws.cell(row=8, column=1).value == "Canopy"
    # El peligro faltante de esa fila queda como [PENDIENTE].
    assert "[PENDIENTE:" in str(ws.cell(row=8, column=2).value)


# --- Flujo completo por el endpoint (motor xlsx) -----------------------------

MATRIX_CHECKLIST = [
    {"field_key": "company_name", "description": "Razón social", "required": True},
    {"field_key": "scope", "description": "Alcance", "required": True},
    {"field_key": "methodology", "description": "Metodología", "required": True},
    {"field_key": "risks", "description": "Filas", "required": True},
    {"field_key": "opportunities", "description": "Oportunidades", "required": False},
]

MATRIX_OUTPUT = {
    "company_name": "Aventura Andina SAS",
    "scope": "Rafting",
    "methodology": "3x3",
    "risks": [{"activity": "Rafting", "hazard": "Volcamiento"}],
    "opportunities": ["Certificar guías"],
}


class _Provider:
    async def embed(self, text: str) -> list[float]:
        return [0.0] * 768

    async def generate_json(self, prompt: str) -> dict:
        return MATRIX_OUTPUT


def test_matrix_endpoint_uses_xlsx_engine(client, make_token, monkeypatch):
    captured: dict = {}
    monkeypatch.setattr(service, "get_ai_provider", lambda: _Provider())
    monkeypatch.setattr(
        repository, "get_active_prompt",
        lambda dt, token: {"version": "v1", "content": "<<JSON_SCHEMA>>"},
    )
    monkeypatch.setattr(repository, "get_company", lambda tid, token: {"name": "X"})
    monkeypatch.setattr(repository, "get_onboarding_data", lambda tid, token: {})
    monkeypatch.setattr(repository, "get_checklist", lambda dt, token: MATRIX_CHECKLIST)
    monkeypatch.setattr(repository, "get_generation_patterns", lambda dt, token: [])

    # Devuelve el MISMO chunk para cada numeral consultado: prueba a la vez
    # que la matriz consulta 6.1.1 y el Anexo A, y que el service deduplica.
    def fake_match(emb, numeral, token, **kw):
        captured.setdefault("numerales", []).append(numeral)
        return [{
            "id": "cccccccc-0000-0000-0000-000000000001",
            "source": "acotur", "numeral": "6.1.1", "content": "Evidencias...",
        }]

    monkeypatch.setattr(repository, "match_knowledge_chunks", fake_match)
    async def fake_list_items(tid, token, category=None):
        return []

    monkeypatch.setattr(inventory_service, "list_items", fake_list_items)
    monkeypatch.setattr(repository, "next_version", lambda tid, dt, token: 1)

    def fake_upload(tid, dt, v, content, token, engine="docx"):
        captured["engine"] = engine
        return f"{tid}/{dt}/v{v}.{engine}"

    monkeypatch.setattr(repository, "upload_document", fake_upload)

    def fake_insert(record, token):
        captured["record"] = record
        return {"id": "doc-x", **record}

    monkeypatch.setattr(repository, "insert_document", fake_insert)
    monkeypatch.setattr(
        repository, "upsert_document_status", lambda *a, **k: None
    )

    token = make_token(tenant_id=TENANT_A, role="empresa")
    resp = client.post(
        "/generation/matriz_riesgos",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert captured["engine"] == "xlsx"
    assert body["storage_path"].endswith("v1.xlsx")
    assert body["template_version"] == "matriz-riesgos-xlsx-v2"
    assert body["completeness"] == 100.0
    # RAG multi-numeral: la matriz consulta 6.1.1 Y el Anexo A…
    assert captured["numerales"] == ["6.1.1", "A"]
    # …y el chunk repetido entre numerales se deduplica en el snapshot.
    assert captured["record"]["rag_chunks_ids"] == [
        "cccccccc-0000-0000-0000-000000000001"
    ]


def test_specs_declare_rag_numerales():
    from app.modules.generation.generators.factory import get_spec

    assert get_spec("matriz_riesgos").rag_numerales == ("6.1.1", "A")
    assert get_spec("politica_seguridad").rag_numerales == ("5.2",)
