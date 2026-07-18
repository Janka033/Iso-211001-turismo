"""Tests de la Matriz de requisitos legales (6.1.3): generador 8/8.

El catálogo legal (MT-03 de Felipe) es fijo y transcrito exacto; la IA solo
aporta identidad y actividades. El filtro por actividad es determinista.
"""

import io

from openpyxl import load_workbook

from app.modules.generation.generators.factory import get_spec
from app.modules.generation.generators.requisitos_legales import (
    _LEGAL_BASE,
    RequisitosLegalesGenerator,
    applicable_norms,
)
from app.modules.generation.schemas import MatrizRequisitosLegalesVariables


def _cells(content: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(content))
    values: list[str] = []
    for ws in wb:
        for row in ws.iter_rows():
            values.extend(str(c.value) for c in row if c.value is not None)
    return values


def test_catalogo_completo_y_citas_exactas():
    """El catálogo tiene las 23 normas de Felipe con citas clave intactas."""
    assert len(_LEGAL_BASE) == 23
    numeros = {n.numero for n in _LEGAL_BASE}
    # Muestras de citas críticas transcritas del MT-03.
    assert "300 de 1996" in numeros  # Ley General de Turismo (RNT)
    assert "1072 de 2015" in numeros  # Decreto Único del Sector Trabajo
    assert "1581 de 2012" in numeros  # Protección de datos
    assert "3840 de 2009" in numeros  # Código de conducta ESCNNA


def test_filtro_por_actividad_es_determinista():
    # La Resolución 3124 (cuatrimotos) solo aplica a tenants con ATV.
    rafting = {n.numero for n in applicable_norms(["rafting"])}
    atv = {n.numero for n in applicable_norms(["atv", "rafting"])}
    assert "3124 de 2014" not in rafting
    assert "3124 de 2014" in atv
    # Con ATV aplican las 23 filas (la Ley 9 de 1979 aparece dos veces:
    # incendios y SST — por eso se cuenta la lista, no el set).
    assert len(applicable_norms(["atv", "rafting"])) == 23
    assert len(applicable_norms(["rafting"])) == 22


def test_matriz_full_document():
    variables = MatrizRequisitosLegalesVariables(
        company_name="Rafting Fonce SAS", activities=["rafting"]
    )
    result = RequisitosLegalesGenerator().generate(variables)

    assert result.pending_fields == []
    txt = "\n".join(_cells(result.content))
    assert "MATRIZ DE REQUISITOS LEGALES" in txt
    assert "Ley" in txt and "300 de 1996" in txt
    assert "3124 de 2014" not in txt  # sin ATV no aplica cuatrimotos
    assert "(se diligencia en el seguimiento)" in txt
    assert "numeral 6.1.3" in txt


def test_matriz_empty_marks_pending_but_catalog_stays():
    result = RequisitosLegalesGenerator().generate(MatrizRequisitosLegalesVariables())
    assert "company_name" in result.pending_fields
    txt = "\n".join(_cells(result.content))
    # El catálogo sale completo aunque falte la identidad.
    assert "300 de 1996" in txt


def test_registered_in_factory():
    spec = get_spec("matriz_requisitos_legales")
    assert spec is not None
    assert spec.numeral == "6.1.3"
    assert spec.variables_model is MatrizRequisitosLegalesVariables
    assert spec.generator.engine == "xlsx"
    assert spec.generator.static_sections
